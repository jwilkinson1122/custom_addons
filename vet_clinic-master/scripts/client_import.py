#!/usr/bin/env python
# -*- coding: utf-8 -*-
# =============================================================================
# For copyright and license notices, see __openerp__.py file in root directory
# =============================================================================

from __future__ import unicode_literals, print_function
import jsonrpclib
from openpyxl import load_workbook


HOST = 'localhost'
PORT = '8069'
DATABASE = 'F2016'
USERNAME = 'admin'
PASSWORD = 'admin'

EXCEL_FILE_PATH = 'clients.xlsx'


# class OdooRpc(object):
#     def __init__(self):
#         common = xmlrpclib.ServerProxy('{}/xmlrpc/common'.format(URL))
#         self.uid = common.login(DATABASE, USERNAME, PASSWORD)
#
#         self.models = xmlrpclib.ServerProxy('{}/xmlrpc/object'.format(URL))
#
#         self.db = DATABASE
#         self.password = PASSWORD
#
#     def execute(self, model, function, args=None, context=None):
#         """
#         :param model: model
#         :param function: a dictionary to call
#         :param args: a list of arguments
#         :param context: dictionary
#         :return:
#         """
#         return self.models.execute(self.db, self.uid, self.password,
#                                       model, function,
#                                       args or [], context or {})


class OdooRpc(object):
    def __init__(self, host, port, db_name, user, password):
        # server proxy object
        url = "http://{host}:{port}/jsonrpc".format(host=host, port=port)
        self.server = jsonrpclib.Server(url)

        # log in the given database
        self.uid = self.server.call(service="common", method="login", args=[db_name, user, password])
        self.password = password
        self.db_name = db_name

    # helper function for invoking model methods
    def execute(self, model, method, *args):
        args = [self.db_name, self.uid, self.password, model, method] + list(args)
        return self.server.call(service="object", method="execute", args=args)


class ImportExcel(object):
    def __init__(self, excel_file):
        wb = load_workbook(excel_file, read_only=True)
        self.ws = wb.active

        self.header = 1

        self.current_row = 0
        self.partner = {
            'animals': []
        }

        self.partner_header = (
            'name',
            'phone'
        )
        self.animal_header = (
            '',
            'name',
            'type',
            'sex',
            'age'
        )

        self.odoo = OdooRpc(HOST, PORT, DATABASE, USERNAME, PASSWORD)

    def process_ws(self):
        for row_counter, row in enumerate(self.ws.iter_rows()):
            print(row_counter)
            if row_counter > self.header - 1:
                if row:
                    self.process_row(row)
                    print()
        if self.partner:
            self.create_partner(self.partner)

    def process_row(self, row):
        animal = {}
        for column_counter, field in enumerate(row):
            print(field.value, sep=' : ', end='')

            if column_counter == 0:
                if field.value:
                    partner = True
                    if self.partner and self.partner.get('name'):
                        self.create_partner(self.partner)
                    self.partner = {
                        'animals': []
                    }
                else:
                    partner = False

            if field.value:
                if partner:
                    self.partner[self.partner_header[column_counter]] = field.value
                else:
                    animal[self.animal_header[column_counter]] = field.value

        if animal:
            self.partner['animals'].append(animal)

    def create_partner(self, values):
        self.odoo.execute('res.partner', 'rpc_create2', values)


if __name__ == '__main__':
    excel = ImportExcel(EXCEL_FILE_PATH)
    excel.process_ws()
